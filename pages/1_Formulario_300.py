"""
=============================================================================
  Extractor de Formulario 300 (IVA)
=============================================================================
"""

import re
import io
import streamlit as st
import pdfplumber
import pandas as pd
from collections import OrderedDict

# ─────────────────────────────────────────────────────────────────────────────
# CONSTANTES F300
# ─────────────────────────────────────────────────────────────────────────────

PERIODO_MAP_BIMESTRAL = {
    "1": "Ene-Feb", "2": "Mar-Abr", "3": "May-Jun",
    "4": "Jul-Ago", "5": "Sep-Oct", "6": "Nov-Dic",
}
PERIODO_MAP_CUATRIMESTRAL = {
    "1": "Ene-Abr", "2": "May-Ago", "3": "Sep-Dic",
}

F300_CONCEPTOS = OrderedDict([
    # Sección: Ingresos
    ("27",  "Ingresos gravados al 5%"),
    ("28",  "Ingresos gravados tarifa general"),
    ("29",  "AIU operaciones gravadas (base especial)"),
    ("30",  "Exportación de bienes"),
    ("31",  "Exportación de servicios"),
    ("32",  "Ventas a Soc. Comercialización Internacional"),
    ("33",  "Ventas a Zonas Francas"),
    ("34",  "Juegos de suerte y azar"),
    ("35",  "Operaciones exentas"),
    ("36",  "Venta licores, aperitivos, vinos y similares"),
    ("37",  "Venta gaseosas y similares"),
    ("38",  "Otros ingresos"),
    ("39",  "Operaciones excluidas"),
    ("40",  "Operaciones no gravadas"),
    ("41",  "Total ingresos brutos"),
    ("42",  "Devoluciones en ventas anuladas/rescindidas"),
    ("43",  "Total ingresos netos del período"),
    # Sección: Compras e importaciones
    ("44",  "Compras bienes gravados tarifa 5%"),
    ("45",  "Compras bienes gravados tarifa general"),
    ("46",  "Bienes/servicios gravados Zonas Francas"),
    ("47",  "Bienes no gravados"),
    ("48",  "Bienes excluidos/exentos Zonas Francas"),
    ("49",  "Servicios"),
    ("50",  "Importaciones bienes gravados tarifa 5%"),
    ("51",  "Importaciones bienes gravados tarifa general"),
    ("52",  "Importaciones servicios gravados tarifa 5%"),
    ("53",  "Importaciones servicios gravados tarifa general"),
    ("54",  "Bienes/servicios excluidos, exentos y no gravados"),
    ("55",  "Total compras e importaciones brutas"),
    ("56",  "Devoluciones compras anuladas/rescindidas"),
    ("57",  "Total compras netas del período"),
    # Sección: Impuesto generado
    ("58",  "IVA generado tarifa 5%"),
    ("59",  "IVA generado tarifa general"),
    ("60",  "IVA sobre AIU (base gravable especial)"),
    ("61",  "IVA juegos de suerte y azar"),
    ("62",  "IVA venta cerveza producción nal. o importada"),
    ("63",  "IVA venta gaseosas y similares"),
    ("64",  "IVA venta licores, aperitivos, vinos y similares"),
    ("65",  "IVA retiro inventario (activos fijos/consumo/donaciones)"),
    ("66",  "IVA recuperado devoluciones compras"),
    ("67",  "Total impuesto generado operaciones gravadas"),
    # Sección: Impuesto descontable
    ("68",  "Imp. descontable importaciones tarifa 5%"),
    ("69",  "Imp. descontable importaciones tarifa general"),
    ("70",  "Imp. desc. bienes/servicios Zonas Francas"),
    ("71",  "Imp. desc. compras bienes tarifa 5%"),
    ("72",  "Imp. desc. compras bienes tarifa general"),
    ("73",  "Imp. desc. licores, aperitivos, vinos y sim."),
    ("74",  "Imp. desc. servicios tarifa 5%"),
    ("75",  "Imp. desc. servicios tarifa general"),
    ("76",  "Descuento IVA exploración hidrocarburos"),
    ("77",  "Total Impuesto pagado o facturado"),
    ("78",  "IVA retenido servicios no residentes"),
    ("79",  "IVA por devoluciones ventas anuladas/rescindidas"),
    ("80",  "Ajuste impuestos descontables (pérdidas/hurto)"),
    ("81",  "Total impuestos descontables"),
    # Sección: Liquidación privada
    ("82",  "Saldo a pagar por el período fiscal"),
    ("83",  "Saldo a favor del período fiscal"),
    ("84",  "Saldo a favor del período fiscal anterior"),
    ("85",  "Retenciones por IVA que le practicaron"),
    ("86",  "Saldo a pagar por impuesto"),
    ("87",  "Sanciones"),
    ("88",  "Total saldo a pagar"),
    ("89",  "Total saldo a favor"),
    # Sección: Control de saldos
    ("90",  "Saldo a favor susceptible devolución/compensación"),
    ("91",  "Saldo a favor susceptible imputar período sig."),
    ("92",  "Saldo a favor sin derecho a dev./comp."),
    ("93",  "Total saldo a favor a imputar al período"),
    ("100", "Total anticipos IVA Régimen SIMPLE"),
])

def limpiar_valor(texto: str) -> float:
    if not texto or not texto.strip():
        return 0.0
    v = texto.strip().replace('$', '').replace(' ', '')
    negativo = False
    if v.startswith('(') and v.endswith(')'):
        negativo = True
        v = v[1:-1]
    elif v.startswith('-'):
        negativo = True
        v = v[1:]
    v = v.replace(',', '')
    try:
        resultado = float(v)
    except ValueError:
        resultado = 0.0
    return -resultado if negativo else resultado

def extraer_cabecera_300(words: list) -> dict:
    año = None
    periodo = None
    periodicidad = None
    razon_social = None

    year_digits = []
    for w in words:
        if 60 <= w["top"] <= 70 and w["x0"] < 130 and w["text"].strip().isdigit():
            year_digits.append(w)
    
    if len(year_digits) >= 4:
        year_digits.sort(key=lambda w: w["x0"])
        año = "".join(w["text"].strip() for w in year_digits[:4])
    
    for w in words:
        if 60 <= w["top"] <= 70 and 260 <= w["x0"] <= 290 and w["text"].strip().isdigit():
            periodo = w["text"].strip()
            break
    
    for w in words:
        t = w["text"].strip().lower()
        if "bimestral" in t:
            periodicidad = "Bimestral"
            break
        elif "cuatrimestral" in t:
            periodicidad = "Cuatrimestral"
            break
    
    for w in words:
        if 188 <= w["top"] <= 198 and len(w["text"].strip()) > 5:
            razon_social = w["text"].strip()
            break
    
    if periodicidad == "Cuatrimestral":
        periodo_nombre = PERIODO_MAP_CUATRIMESTRAL.get(periodo, f"P{periodo}")
    elif periodicidad == "Bimestral":
        periodo_nombre = PERIODO_MAP_BIMESTRAL.get(periodo, f"P{periodo}")
    else:
        periodo_nombre = PERIODO_MAP_BIMESTRAL.get(periodo, f"P{periodo}")
    
    columna = f"{periodo_nombre} {año}" if año else f"P{periodo}"
    
    return {
        "año": año or "XXXX",
        "periodo": periodo or "??",
        "periodo_nombre": periodo_nombre,
        "periodicidad": periodicidad or "Desconocida",
        "columna": columna,
        "razon_social": razon_social or "Desconocida",
    }

def extraer_valores_300(words: list) -> dict:
    valor_words = []
    for w in words:
        text = w["text"].strip()
        if text and re.match(r'^[\d,.\-()]+$', text) and any(c.isdigit() for c in text):
            valor_words.append({
                "x0": w["x0"],
                "top": w["top"],
                "text": text,
            })
    
    COL_IZQ = (250, 310)
    COL_DER = (540, 595)
    COL_MID = (440, 470)
    
    izq_vals = []
    der_vals = []
    mid_vals = []
    
    for w in valor_words:
        if 210 <= w["top"] <= 650:
            if COL_IZQ[0] <= w["x0"] <= COL_IZQ[1]:
                izq_vals.append((round(w["top"]), limpiar_valor(w["text"])))
            elif COL_DER[0] <= w["x0"] <= COL_DER[1]:
                der_vals.append((round(w["top"]), limpiar_valor(w["text"])))
            elif COL_MID[0] <= w["x0"] <= COL_MID[1]:
                mid_vals.append((round(w["top"]), limpiar_valor(w["text"])))
    
    izq_vals.sort(key=lambda x: x[0])
    der_vals.sort(key=lambda x: x[0])
    mid_vals.sort(key=lambda x: x[0])
    
    casillas_izq = list(range(27, 62))
    casillas_der = list(range(62, 94)) + [100]
    
    resultado = {}
    
    for i, (top, val) in enumerate(izq_vals):
        if i < len(casillas_izq):
            resultado[str(casillas_izq[i])] = val
    
    for i, (top, val) in enumerate(der_vals):
        if i < len(casillas_der):
            resultado[str(casillas_der[i])] = val
            
    return resultado

def procesar_pdf_300(pdf_file) -> dict:
    resultado = {
        "nombre_archivo": pdf_file.name,
        "cabecera": None,
        "casillas": {},
        "error": None,
    }
    try:
        with pdfplumber.open(pdf_file) as pdf:
            if len(pdf.pages) == 0:
                resultado["error"] = f"⚠️ '{pdf_file.name}' no tiene páginas."
                return resultado
            page = pdf.pages[0]
            words = page.extract_words(keep_blank_chars=True, x_tolerance=3, y_tolerance=3)
            
            if not words:
                resultado["error"] = f"⚠️ '{pdf_file.name}' no contiene texto extraíble."
                return resultado
            
            resultado["cabecera"] = extraer_cabecera_300(words)
            resultado["casillas"] = extraer_valores_300(words)
    except Exception as e:
        resultado["error"] = f"❌ Error procesando '{pdf_file.name}': {str(e)}"
    return resultado

def construir_df_300(resultados: list) -> pd.DataFrame:
    exitosos = [r for r in resultados if r["casillas"] and r["cabecera"]]
    if not exitosos:
        return pd.DataFrame()
    
    index_tuples = [(cas, concepto) for cas, concepto in F300_CONCEPTOS.items()]
    index = pd.MultiIndex.from_tuples(index_tuples, names=["Casilla", "Concepto"])
    
    periodos = []
    for r in exitosos:
        col = r["cabecera"]["columna"]
        if col not in periodos:
            periodos.append(col)
    
    df = pd.DataFrame(index=index, columns=periodos, dtype=float)
    df[:] = 0.0
    
    for r in exitosos:
        periodo = r["cabecera"]["columna"]
        for casilla, valor in r["casillas"].items():
            if casilla in F300_CONCEPTOS:
                concepto = F300_CONCEPTOS[casilla]
                df.loc[(casilla, concepto), periodo] = valor
    
    return df

def df_a_excel(df: pd.DataFrame) -> bytes:
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        sheet_name = "F300 Consolidado"
        df.to_excel(writer, sheet_name=sheet_name, merge_cells=True)
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        ws = writer.sheets[sheet_name]
        
        header_font = Font(bold=True, color="FFFFFF", size=10, name="Calibri")
        header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        thin_border = Border(
            left=Side(style='thin', color='B4C6E7'), right=Side(style='thin', color='B4C6E7'),
            top=Side(style='thin', color='B4C6E7'), bottom=Side(style='thin', color='B4C6E7'),
        )
        
        for row_idx in range(1, 4):
            for cell in ws[row_idx]:
                if row_idx <= 1:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_align
                cell.border = thin_border
        
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.border = thin_border
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '#,##0'
                    cell.alignment = Alignment(horizontal="right")
        
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 60
        for col_idx in range(3, ws.max_column + 1):
            from openpyxl.utils import get_column_letter
            ws.column_dimensions[get_column_letter(col_idx)].width = 18
            
        ws.freeze_panes = "C2"
    return buffer.getvalue()

# ─────────────────────────────────────────────────────────────────────────────
# UI STREAMLIT
# ─────────────────────────────────────────────────────────────────────────────

st.set_page_config(page_title="F300 - IVA", page_icon="📝", layout="wide")

st.markdown("""
<style>
    .metric-card {
        background: linear-gradient(145deg, #1e293b, #0f172a);
        padding: 1.5rem; border-radius: 12px;
        border: 1px solid rgba(99, 102, 241, 0.2);
        text-align: center; color:white;
    }
    .metric-value { font-size: 2.2rem; font-weight: 700; color: #60a5fa; }
    .status-success { background: #064e3b; color: #6ee7b7; padding: 0.6rem; border-radius: 8px; margin: 5px 0;}
    .status-error { background: #7f1d1d; color: #fca5a5; padding: 0.6rem; border-radius: 8px; margin: 5px 0;}
    .subtle-divider { border-top: 1px solid rgba(255,255,255,0.1); margin: 1.5rem 0;}
</style>
""", unsafe_allow_html=True)

st.title("📝 Módulo F300 - IVA")
st.markdown("Extrae y consolida datos de múltiples PDFs del formulario 300 de la DIAN (formato unificado).")

archivos_pdf = st.file_uploader(
    "📁 Sube los PDFs del Formulario 300 (IVA)",
    type=["pdf"],
    accept_multiple_files=True
)

if archivos_pdf:
    st.markdown('<div class="subtle-divider"></div>', unsafe_allow_html=True)
    st.subheader("🔄 Procesando archivos...")
    progress = st.progress(0)
    resultados = []
    errores = []

    for i, archivo in enumerate(archivos_pdf):
        progress.progress((i + 1) / len(archivos_pdf))
        res = procesar_pdf_300(archivo)
        resultados.append(res)
        if res.get("error"):
            errores.append(res)
    progress.empty()

    exitosos = [r for r in resultados if r.get("casillas")]
    col1, col2, col3 = st.columns(3)
    with col1:
        st.markdown(f'<div class="metric-card"><div class="metric-value">{len(archivos_pdf)}</div>PDFs Cargados</div>', unsafe_allow_html=True)
    with col2:
        st.markdown(f'<div class="metric-card"><div class="metric-value">{len(exitosos)}</div>Procesados OK</div>', unsafe_allow_html=True)
    with col3:
        st.markdown(f'<div class="metric-card"><div class="metric-value">{len(errores)}</div>Con Errores</div>', unsafe_allow_html=True)

    with st.expander("📄 Detalle por Archivo", expanded=bool(errores)):
        for r in resultados:
            if r.get("error"):
                st.markdown(f'<div class="status-error">{r["error"]}</div>', unsafe_allow_html=True)
            else:
                cab = r.get("cabecera", {})
                st.markdown(f'<div class="status-success">✅ <strong>{r["nombre_archivo"]}</strong> → {cab.get("columna", "?")}</div>', unsafe_allow_html=True)

    if exitosos:
        st.markdown('<div class="subtle-divider"></div>', unsafe_allow_html=True)
        st.subheader("📊 Datos Consolidados")
        df = construir_df_300(resultados)
        
        if df is not None and not df.empty:
            st.dataframe(df.style.format("{:,.0f}", na_rep="—"), use_container_width=True, height=700)
            excel_bytes = df_a_excel(df)
            st.download_button(
                label="📥 Descargar Excel",
                data=excel_bytes,
                file_name="DIAN_F300_Consolidado.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                help="Descarga los datos compilados listos para Excel."
            )
else:
    st.info("Sube uno o más PDFs en el recuadro superior para empezar.")
