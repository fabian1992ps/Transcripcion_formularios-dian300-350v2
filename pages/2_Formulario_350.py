"""
=============================================================================
  Extractor de Formulario 350 (Retención en la Fuente)
=============================================================================
"""

import re
import io
import streamlit as st
import pdfplumber
import pandas as pd
from collections import OrderedDict

# ─────────────────────────────────────────────────────────────────────────────
# CONSTANTES F350
# ─────────────────────────────────────────────────────────────────────────────

PERIODO_MAP_MENSUAL = {
    "1": "Enero", "2": "Febrero", "3": "Marzo", "4": "Abril",
    "5": "Mayo", "6": "Junio", "7": "Julio", "8": "Agosto",
    "9": "Septiembre", "10": "Octubre", "11": "Noviembre", "12": "Diciembre",
}

F350_CONCEPTOS_RETENCION = [
    ("Rentas de trabajo",                      None, None, None, "77",  None, None, None, "93"),
    ("Rentas de pensiones",                    None, None, None, "78",  None, None, None, "94"),
    ("Honorarios",                             "29", "42", None, None,  "79", "95", None, None),
    ("Comisiones",                             "30", "43", None, None,  "80", "96", None, None),
    ("Servicios",                              "31", "44", None, None,  "81", "97", None, None),
    ("Rendimientos financieros e intereses",   "32", "45", None, None,  "82", "98", None, None),
    ("Arrendamientos (Muebles e inmuebles)",   "33", "46", None, None,  "83", "99", None, None),
    ("Regalías y explotación propiedad int.",  "34", "47", None, None,  "84", "100", None, None),
    ("Dividendos y participaciones",           "35", "48", None, None,  "85", "101", None, None),
    ("Compras",                                "36", "49", None, None,  "86", "102", None, None),
    ("Transacciones tarjetas débito/crédito",  "37", "50", None, None,  "87", "103", None, None),
    ("Contratos de construcción",              "38", "51", None, None,  "88", "104", None, None),
    ("Enajenación activos fijos PN",           None, None, None, None,  "89", "105", None, None),
    ("Loterías, rifas, apuestas y similares",  "39", "52", None, None,  "90", "106", None, None),
    ("Hidrocarburos, carbón y prod. mineros",  "40", "53", None, None,  "91", "107", None, None),
    ("Otros pagos sujetos a retención",        "41", "54", None, None,  "92", "108", None, None),
    ("Pagos al exterior sin convenio",         "55", "57", None, None,  "109", "111", None, None),
    ("Pagos al exterior con convenio",         "56", "58", None, None,  "110", "112", None, None),
]

F350_CONCEPTOS_AUTORRETENCION = [
    ("Autorretencion - Exonerados aportes",    "59", "68", None, None,  None, None, None, None),
    ("Autorretencion - Ventas",                "60", "69", None, None,  "113", "121", None, None),
    ("Autorretencion - Honorarios",            "61", "70", None, None,  "114", "122", None, None),
    ("Autorretencion - Comisiones",            "62", "71", None, None,  "115", "123", None, None),
    ("Autorretencion - Servicios",             "63", "72", None, None,  "116", "124", None, None),
    ("Autorretencion - Rend. financieros",     "64", "73", None, None,  "117", "125", None, None),
    ("Autorretencion - (casilla 65)",          "65", "74", None, None,  "118", "126", None, None),
    ("Autorretencion - (casilla 66)",          "66", "75", None, None,  "119", "127", None, None),
    ("Autorretencion - Otros conceptos",       "67", "76", None, None,  "120", "128", None, None),
]

F350_CONCEPTOS_TOTALES = [
    ("Menos ret. en exceso/indebidas (Renta)", "129"),
    ("Total retenciones renta y complem.",      "130"),
    ("Ret. IVA a responsables",                "131"),
    ("Ret. IVA servicios no residentes",       "132"),
    ("Menos ret. IVA en exceso/indebidas",     "133"),
    ("Total retenciones IVA",                  "134"),
    ("Ret. impuesto timbre nacional",          "135"),
    ("Total retenciones",                      "136"),
    ("Sanciones",                              "137"),
    ("Total retenciones más sanciones",        "138"),
]

def limpiar_valor(texto: str) -> float:
    if not texto or not texto.strip(): return 0.0
    v = texto.strip().replace('$', '').replace(' ', '')
    negativo = False
    if v.startswith('(') and v.endswith(')'):
        negativo = True; v = v[1:-1]
    elif v.startswith('-'):
        negativo = True; v = v[1:]
    v = v.replace(',', '')
    try: return -float(v) if negativo else float(v)
    except ValueError: return 0.0

def extraer_cabecera_350(words: list) -> dict:
    año = None
    periodo = None
    razon_social = None

    year_digits = []
    for w in words:
        if 60 <= w["top"] <= 70 and w["x0"] < 130 and w["text"].strip().isdigit():
            year_digits.append(w)
    
    if len(year_digits) >= 4:
        year_digits.sort(key=lambda w: w["x0"])
        año = "".join(w["text"].strip() for w in year_digits[:4])
    
    # Busqueda del periodo - FIX aplicado (Ampliación y unión de dígitos para meses de dos cifras)
    period_digits = []
    for w in words:
        if 60 <= w["top"] <= 70 and 240 <= w["x0"] <= 330 and w["text"].strip().isdigit():
            period_digits.append(w)
    
    if period_digits:
        period_digits.sort(key=lambda w: w["x0"])
        joined = "".join(w["text"].strip() for w in period_digits)
        try:
            periodo = str(int(joined)) # Convierte '01' o '0' y '1' a '1', y '10' a '10'
        except:
            periodo = joined
            
    for w in words:
        if 188 <= w["top"] <= 198 and len(w["text"].strip()) > 5:
            razon_social = w["text"].strip()
            break
    
    periodo_nombre = PERIODO_MAP_MENSUAL.get(periodo, f"P{periodo}" if periodo else "??")
    columna = f"{periodo_nombre} {año}" if año else f"P{periodo}"
    
    return {
        "año": año or "XXXX",
        "periodo": periodo or "??",
        "periodo_nombre": periodo_nombre,
        "columna": columna,
        "razon_social": razon_social or "Desconocida",
    }

def extraer_valores_350(words: list) -> dict:
    valor_words = []
    for w in words:
        text = w["text"].strip()
        if text and re.match(r'^[\d,.\-()]+$', text) and any(c.isdigit() for c in text):
            valor_words.append({"x0": w["x0"], "top": w["top"], "text": text})
    
    COL_RANGES = {
        "col1": (200, 260), "col2": (320, 370), "col3": (430, 480), "col4": (545, 595),
    }
    def clasificar_columna(x0):
        for col_name, (xmin, xmax) in COL_RANGES.items():
            if xmin <= x0 <= xmax: return col_name
        return None
    
    filas_valores = {}
    for w in valor_words:
        if w["top"] < 220 or w["top"] > 700: continue
        col = clasificar_columna(w["x0"])
        if col is None: continue
        top_key = round(w["top"])
        if top_key not in filas_valores: filas_valores[top_key] = {}
        filas_valores[top_key][col] = limpiar_valor(w["text"])
    
    return filas_valores

def procesar_pdf_350(pdf_file) -> dict:
    resultado = {
        "nombre_archivo": pdf_file.name,
        "cabecera": None,
        "filas_valores": {},
        "error": None,
    }
    try:
        with pdfplumber.open(pdf_file) as pdf:
            if len(pdf.pages) == 0:
                resultado["error"] = f"⚠️ '{pdf_file.name}' no tiene páginas."
                return resultado
            page = pdf.pages[0]
            words = page.extract_words(keep_blank_chars=True, x_tolerance=3, y_tolerance=3)
            
            if not words:
                resultado["error"] = f"⚠️ '{pdf_file.name}' no contiene texto extraíble."
                return resultado
            
            resultado["cabecera"] = extraer_cabecera_350(words)
            resultado["filas_valores"] = extraer_valores_350(words)
            
            if not resultado["filas_valores"]:
                resultado["error"] = f"⚠️ No se encontraron valores en '{pdf_file.name}'."
    except Exception as e:
        resultado["error"] = f"❌ Error procesando '{pdf_file.name}': {str(e)}"
    return resultado

def construir_df_350(resultados: list) -> pd.DataFrame:
    exitosos = [r for r in resultados if r["filas_valores"] and r["cabecera"]]
    if not exitosos: return pd.DataFrame()
    
    conceptos_renta = [
        "Rentas de trabajo", "Rentas de pensiones", "Honorarios", "Comisiones", "Servicios",
        "Rendimientos financieros e intereses", "Arrendamientos (Muebles e inmuebles)",
        "Regalías y explotación propiedad int.", "Dividendos y participaciones", "Compras",
        "Transacciones tarjetas débito/crédito", "Contratos de construcción",
        "Enajenación activos fijos PN", "Loterías, rifas, apuestas y similares",
        "Hidrocarburos, carbón y prod. mineros", "Otros pagos sujetos a retención",
        "Pagos al exterior sin convenio", "Pagos al exterior con convenio",
    ]
    conceptos_autor = [
        "Autorretencion - Exonerados aportes", "Autorretencion - Ventas",
        "Autorretencion - Honorarios", "Autorretencion - Comisiones",
        "Autorretencion - Servicios", "Autorretencion - Rend. financieros",
        "Autorretencion - (casilla 65)", "Autorretencion - (casilla 66)", "Autorretencion - Otros conceptos",
    ]
    conceptos_totales = [c[0] for c in F350_CONCEPTOS_TOTALES]
    conceptos_orden = conceptos_renta + conceptos_autor + conceptos_totales
    
    periodos = []
    for r in exitosos:
        col = r["cabecera"]["columna"]
        if col not in periodos: periodos.append(col)
    
    col_tuples = [(p, s) for p in periodos for s in ["Base PJ", "Ret PJ", "Base PN", "Ret PN"]]
    multi_cols = pd.MultiIndex.from_tuples(col_tuples, names=["Período", "Tipo"])
    
    df = pd.DataFrame(index=conceptos_orden, columns=multi_cols, dtype=float)
    df.index.name = "Concepto"
    df[:] = 0.0
    
    for r in exitosos:
        periodo = r["cabecera"]["columna"]
        filas = r["filas_valores"]
        tops_ordenados = sorted(filas.keys())
        
        tops_con_multicol, tops_solo_col4 = [], []
        for t in tops_ordenados:
            cols = filas[t]
            if "col1" in cols or "col2" in cols or "col3" in cols: tops_con_multicol.append(t)
            elif "col4" in cols and len(cols) == 1: tops_solo_col4.append(t)
            else: tops_con_multicol.append(t)
        
        todos_conceptos_multi = conceptos_renta + conceptos_autor
        for i, top_y in enumerate(tops_con_multicol):
            if i >= len(todos_conceptos_multi): break
            concepto = todos_conceptos_multi[i]
            vals = filas[top_y]
            if "col1" in vals: df.loc[concepto, (periodo, "Base PJ")] = vals["col1"]
            if "col2" in vals: df.loc[concepto, (periodo, "Ret PJ")] = vals["col2"]
            if "col3" in vals: df.loc[concepto, (periodo, "Base PN")] = vals["col3"]
            if "col4" in vals: df.loc[concepto, (periodo, "Ret PN")] = vals["col4"]
        
        for i, top_y in enumerate(tops_solo_col4):
            if i >= len(conceptos_totales): break
            concepto = conceptos_totales[i]
            vals = filas[top_y]
            if "col4" in vals: df.loc[concepto, (periodo, "Ret PN")] = vals["col4"]
    return df

def df_a_excel(df: pd.DataFrame) -> bytes:
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        sheet_name = "F350 Consolidado"
        df.to_excel(writer, sheet_name=sheet_name, merge_cells=True)
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        ws = writer.sheets[sheet_name]
        
        header_font = Font(bold=True, color="FFFFFF", size=10, name="Calibri")
        header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
        subheader_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        thin_border = Border(
            left=Side(style='thin', color='B4C6E7'), right=Side(style='thin', color='B4C6E7'),
            top=Side(style='thin', color='B4C6E7'), bottom=Side(style='thin', color='B4C6E7'),
        )
        total_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
        total_font = Font(bold=True, size=10, name="Calibri")
        
        for row_idx in range(1, 4):
            for cell in ws[row_idx]:
                if row_idx <= 2:
                    cell.font = header_font
                    cell.fill = header_fill if row_idx == 1 else subheader_fill
                    cell.alignment = header_align
                cell.border = thin_border
        
        for row in ws.iter_rows(min_row=3):
            for cell in row:
                cell.border = thin_border
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '#,##0'
                    cell.alignment = Alignment(horizontal="right")
                
                row_label = ws.cell(row=cell.row, column=1).value
                if row_label and isinstance(row_label, str):
                    if "total" in row_label.lower() or "saldo" in row_label.lower():
                        cell.fill = total_fill
                        cell.font = total_font
        
        ws.column_dimensions['A'].width = 40
        for col_idx in range(2, ws.max_column + 1):
            from openpyxl.utils import get_column_letter
            ws.column_dimensions[get_column_letter(col_idx)].width = 15
            
        ws.freeze_panes = "B3"
    return buffer.getvalue()

# ─────────────────────────────────────────────────────────────────────────────
# UI STREAMLIT
# ─────────────────────────────────────────────────────────────────────────────

st.set_page_config(page_title="F350 - Retención", page_icon="🏢", layout="wide")

st.markdown("""
<style>
    .metric-card {
        background: linear-gradient(145deg, #1e293b, #0f172a);
        padding: 1.5rem; border-radius: 12px;
        border: 1px solid rgba(99, 102, 241, 0.2);
        text-align: center; color:white;
    }
    .metric-value { font-size: 2.2rem; font-weight: 700; color: #a78bfa; }
    .status-success { background: #064e3b; color: #6ee7b7; padding: 0.6rem; border-radius: 8px; margin: 5px 0;}
    .status-error { background: #7f1d1d; color: #fca5a5; padding: 0.6rem; border-radius: 8px; margin: 5px 0;}
    .subtle-divider { border-top: 1px solid rgba(255,255,255,0.1); margin: 1.5rem 0;}
</style>
""", unsafe_allow_html=True)

st.title("🏢 Módulo F350 - Retención en la Fuente")
st.markdown("Extrae y consolida datos de múltiples PDFs del formulario 350. ¡Incluye corrección para extracciones de meses de dos dígitos (ej. Octubre, Noviembre, Diciembre)!")

archivos_pdf = st.file_uploader(
    "📁 Sube los PDFs del Formulario 350 (Retención en la Fuente)",
    type=["pdf"],
    accept_multiple_files=True
)

if archivos_pdf:
    st.markdown('<div class="subtle-divider"></div>', unsafe_allow_html=True)
    st.subheader("🔄 Procesando archivos...")
    progress = st.progress(0)
    resultados = []
    errores = []

    for i, archivo in enumerate(archivos_pdf):
        progress.progress((i + 1) / len(archivos_pdf))
        res = procesar_pdf_350(archivo)
        resultados.append(res)
        if res.get("error"):
            errores.append(res)
    progress.empty()

    exitosos = [r for r in resultados if r.get("filas_valores")]
    col1, col2, col3 = st.columns(3)
    with col1:
        st.markdown(f'<div class="metric-card"><div class="metric-value">{len(archivos_pdf)}</div>PDFs Cargados</div>', unsafe_allow_html=True)
    with col2:
        st.markdown(f'<div class="metric-card"><div class="metric-value">{len(exitosos)}</div>Procesados OK</div>', unsafe_allow_html=True)
    with col3:
        st.markdown(f'<div class="metric-card"><div class="metric-value">{len(errores)}</div>Con Errores</div>', unsafe_allow_html=True)

    with st.expander("📄 Detalle por Archivo", expanded=bool(errores)):
        for r in resultados:
            if r.get("error"):
                st.markdown(f'<div class="status-error">{r["error"]}</div>', unsafe_allow_html=True)
            else:
                cab = r.get("cabecera", {})
                st.markdown(f'<div class="status-success">✅ <strong>{r["nombre_archivo"]}</strong> → {cab.get("columna", "?")} | {len(r.get("filas_valores",{}))} filas</div>', unsafe_allow_html=True)

    if exitosos:
        st.markdown('<div class="subtle-divider"></div>', unsafe_allow_html=True)
        st.subheader("📊 Datos Consolidados")
        df = construir_df_350(resultados)
        
        if df is not None and not df.empty:
            st.dataframe(df.style.format("{:,.0f}", na_rep="—"), use_container_width=True, height=700)
            excel_bytes = df_a_excel(df)
            st.download_button(
                label="📥 Descargar Excel",
                data=excel_bytes,
                file_name="DIAN_F350_Consolidado.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                help="Descarga los datos compilados listos para Excel."
            )
else:
    st.info("Sube uno o más PDFs en el recuadro superior para empezar.")
